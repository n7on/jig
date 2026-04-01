#!/usr/bin/env python3
"""Convert TSV input (stdin) to a formatted Excel (.xlsx) file."""

import argparse
import csv
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def write_sheet(ws, headers, data):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    alt_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, header in enumerate(headers, start=1):
        col_values = [header] + [
            str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
            for row in data
        ]
        max_len = max(len(v) for v in col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description="Convert TSV to Excel")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name")
    args = parser.parse_args()

    reader = csv.reader(sys.stdin, delimiter="\t")
    rows = list(reader)

    if not rows:
        print("No input data", file=sys.stderr)
        sys.exit(1)

    headers = rows[0]
    data = rows[1:]

    if os.path.exists(args.output):
        wb = openpyxl.load_workbook(args.output)
        if args.sheet in wb.sheetnames:
            print(f"Sheet '{args.sheet}' already exists in {args.output}", file=sys.stderr)
            sys.exit(1)
        ws = wb.create_sheet(title=args.sheet)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = args.sheet

    write_sheet(ws, headers, data)
    wb.save(args.output)


if __name__ == "__main__":
    main()
